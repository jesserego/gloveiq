#!/usr/bin/env python3
"""
SidelineSwap Master Scraper — GloveIQ Structured Edition
"""

import argparse
import re
import json
import time
from urllib.parse import urljoin
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

DEFAULT_START_URL = "https://sidelineswap.com/shop/baseball/baseball-gloves/l78"
SHEET_NAME = "Catalog"

def clean(s):
    return re.sub(r"\s+", " ", (s or "").strip())

def extract_size_inches(text):
    t = clean(text)
    if not t:
        return None
    m = re.search(r"(\d{1,2}\.\d{1,2})", t)
    if m:
        return float(m.group(1))
    m = re.search(r"(\d{1,2})\s*(?:-|\s)\s*(\d)\s*/\s*(\d)", t)
    if m:
        whole = float(m.group(1))
        num = float(m.group(2))
        den = float(m.group(3))
        return whole + (num / den)
    return None

def norm_throw(v):
    v = clean(v).lower()
    if "right" in v and "throw" in v:
        return "RHT"
    if "left" in v and "throw" in v:
        return "LHT"
    return "UNK"

def norm_position(v):
    v = clean(v).lower()
    if "outfield" in v or v == "of":
        return "OF"
    if "infield" in v or v == "if":
        return "IF"
    if "pitch" in v:
        return "P"
    if "catch" in v:
        return "C"
    if "first" in v or "1b" in v:
        return "1B"
    if "middle" in v or "ss" in v or "2b" in v:
        return "MI"
    return clean(v).upper() if v else None

def normalize_specs(specs, title):
    lower = { clean(k).lower(): v for k, v in (specs or {}).items() }

    def get(*keys):
        for k in keys:
            if k in lower and lower[k]:
                return lower[k]
        return None

    size_raw = get("size", "glove size")
    throw_raw = get("throwing hand", "throws")
    pos_raw = get("position")
    web_raw = get("web", "web type")

    return {
        "raw": specs,
        "norm": {
            "size_in": extract_size_inches(str(size_raw)) if size_raw else None,
            "throw_hand": norm_throw(str(throw_raw)) if throw_raw else "UNK",
            "position": norm_position(str(pos_raw)) if pos_raw else None,
            "web": clean(str(web_raw)) if web_raw else None,
            "title": clean(str(title)) if title else None,
            "source": "SS"
        }
    }

def fetch(url):
    headers = {"User-Agent": "Mozilla/5.0"}
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    return r.text

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--start-url", default=DEFAULT_START_URL)
    parser.add_argument("--max-pages", type=int, default=5)
    parser.add_argument("--max-details", type=int, default=0)
    parser.add_argument("--delay", type=float, default=1.5)
    args = parser.parse_args()

    wb = load_workbook(args.xlsx)
    ws = wb[SHEET_NAME]

    # Catalog Phase
    current_url = args.start_url
    for page in range(1, args.max_pages + 1):
        print(f"[CATALOG] Page {page}: {current_url}")
        html = fetch(current_url)
        soup = BeautifulSoup(html, "lxml")

        for a in soup.select("a[href*='/gear/']"):
            href = a.get("href")
            if not href:
                continue
            full = urljoin(current_url, href)
            if "sidelineswap.com/gear/" not in full:
                continue
            m = re.search(r"/gear/.+?/(\d+)-", full)
            if not m:
                continue
            listing_id = m.group(1)
            title = clean(a.get_text())
            ws.append([listing_id, full, title])

        wb.save(args.xlsx)

        next_link = soup.find("a", string=re.compile("Next", re.I))
        if not next_link:
            break
        current_url = urljoin(current_url, next_link.get("href"))
        time.sleep(args.delay)

    # Detail Phase
    if args.max_details == 0:
        print("Skipping detail phase.")
        wb.save(args.xlsx)
        return

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        if idx > args.max_details:
            break

        listing_id = row[0].value
        url = row[1].value
        print(f"[DETAIL] {listing_id}")

        html = fetch(url)
        soup = BeautifulSoup(html, "lxml")

        specs = {}
        for li in soup.select("li"):
            if ":" in li.get_text():
                parts = li.get_text().split(":", 1)
                specs[clean(parts[0])] = clean(parts[1])

        title = soup.find("h1").get_text(strip=True) if soup.find("h1") else None

        norm = normalize_specs(specs, title)

        ws.cell(row=idx+1, column=10, value=json.dumps(norm, ensure_ascii=False))

        if idx % 25 == 0:
            wb.save(args.xlsx)

        time.sleep(args.delay)

    wb.save(args.xlsx)
    print("Done.")

if __name__ == "__main__":
    main()
