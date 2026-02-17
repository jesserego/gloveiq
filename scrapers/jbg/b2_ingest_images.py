#!/usr/bin/env python3
"""
GloveIQ Image Ingest (Backblaze B2)
-----------------------------------
Reads an Excel workbook produced by the scrapers and uploads per-listing images to a PRIVATE Backblaze B2 bucket.

Inputs:
- Workbook with detail sheet containing: source, source_listing_id, listing_url, images_json (list of URLs)

Outputs (written back to the same workbook):
- b2_images_json: list of objects with keys: { "b2_key", "file_name", "content_type", "source_url" }
- b2_status: OK | SKIP | ERROR
- b2_error: error message (if any)

Why separate step?
- Scrape can run without credentials.
- Upload can be throttled / resumed safely.

Security:
- Uses environment variables (see .env.example). Bucket remains private.

Requires:
- b2sdk (Backblaze B2 SDK)
- requests
- openpyxl
"""

import argparse
import hashlib
import json
import mimetypes
import os
import time
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import requests
from b2sdk.v2 import InMemoryAccountInfo, B2Api
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_DETAIL_SHEETS = ["SS_Detail_Enrichment", "JBG_Detail_Enrichment"]

def _safe_json_loads(val: Any) -> Any:
    if val is None:
        return None
    if isinstance(val, (dict, list)):
        return val
    s = str(val).strip()
    if not s:
        return None
    try:
        return json.loads(s)
    except Exception:
        return None

def _sha1_bytes(b: bytes) -> str:
    h = hashlib.sha1()
    h.update(b)
    return h.hexdigest()

def _guess_content_type(url: str, fallback: str = "image/jpeg") -> str:
    ct, _ = mimetypes.guess_type(url.split("?")[0])
    return ct or fallback

def _ensure_columns(ws: Worksheet, required: List[str]) -> Dict[str, int]:
    header_row = 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v:
            headers[str(v).strip()] = col

    # Append missing columns at the end
    next_col = ws.max_column + 1
    for name in required:
        if name not in headers:
            ws.cell(row=header_row, column=next_col).value = name
            headers[name] = next_col
            next_col += 1
    return headers

@dataclass
class B2Config:
    key_id: str
    app_key: str
    bucket_name: str
    prefix: str

def _load_b2_config() -> B2Config:
    key_id = os.getenv("B2_KEY_ID", "").strip()
    app_key = os.getenv("B2_APP_KEY", "").strip()
    bucket = os.getenv("B2_BUCKET", "").strip()
    prefix = os.getenv("B2_PREFIX", "gloveiq").strip().strip("/")
    if not key_id or not app_key or not bucket:
        raise RuntimeError("Missing B2 credentials. Set B2_KEY_ID, B2_APP_KEY, B2_BUCKET (and optional B2_PREFIX).")
    return B2Config(key_id=key_id, app_key=app_key, bucket_name=bucket, prefix=prefix)

def _b2_connect(cfg: B2Config):
    info = InMemoryAccountInfo()
    b2 = B2Api(info)
    b2.authorize_account("production", cfg.key_id, cfg.app_key)
    bucket = b2.get_bucket_by_name(cfg.bucket_name)
    return bucket

def _download_image(url: str, timeout: int = 30) -> Tuple[bytes, str]:
    # Some sites need headers
    headers = {
        "User-Agent": "GloveIQBot/1.0 (+https://gloveiq.com; research/scrape)",
        "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": url,
    }
    r = requests.get(url, headers=headers, timeout=timeout)
    r.raise_for_status()
    content_type = r.headers.get("Content-Type", "").split(";")[0].strip()
    if not content_type:
        content_type = _guess_content_type(url)
    return r.content, content_type

def _b2_key(prefix: str, source: str, listing_id: str, index: int, sha1: str, content_type: str) -> str:
    ext = mimetypes.guess_extension(content_type) or ".jpg"
    ext = ext if ext.startswith(".") else f".{ext}"
    # Deterministic key: prefix/source/listing_id/{index}_{sha1[:10]}.ext
    return f"{prefix}/{source}/{listing_id}/{index:02d}_{sha1[:10]}{ext}"

def _row_get(ws: Worksheet, row: int, headers: Dict[str, int], key: str):
    col = headers.get(key)
    return ws.cell(row=row, column=col).value if col else None

def _row_set(ws: Worksheet, row: int, headers: Dict[str, int], key: str, value: Any):
    col = headers[key]
    ws.cell(row=row, column=col).value = value

def ingest_images(
    xlsx_path: str,
    sheets: List[str],
    limit: int,
    delay: float,
    resume: bool,
    dry_run: bool,
):
    cfg = _load_b2_config()
    bucket = _b2_connect(cfg)

    wb = load_workbook(xlsx_path)
    found_any = False

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        found_any = True
        ws = wb[sheet_name]

        headers = _ensure_columns(ws, [
            "source",
            "source_listing_id",
            "listing_url",
            "images_json",
            "b2_images_json",
            "b2_status",
            "b2_error",
            "updated_at",
        ])

        processed = 0
        for r in range(2, ws.max_row + 1):
            if limit and processed >= limit:
                break

            source = str(_row_get(ws, r, headers, "source") or "").strip()
            listing_id = str(_row_get(ws, r, headers, "source_listing_id") or "").strip()
            listing_url = str(_row_get(ws, r, headers, "listing_url") or "").strip()
            images_val = _row_get(ws, r, headers, "images_json")
            b2_existing = _safe_json_loads(_row_get(ws, r, headers, "b2_images_json"))

            if not source or not listing_id or not images_val:
                continue

            if resume and b2_existing:
                _row_set(ws, r, headers, "b2_status", "SKIP")
                continue

            images = _safe_json_loads(images_val)
            if not isinstance(images, list) or not images:
                continue

            out: List[Dict[str, Any]] = []
            try:
                for idx, img_url in enumerate(images, start=1):
                    if not isinstance(img_url, str) or not img_url.strip():
                        continue
                    img_url = img_url.strip()

                    if dry_run:
                        # fake sha1 for deterministic key even in dry-run
                        sha1 = _sha1_bytes(img_url.encode("utf-8"))
                        ct = _guess_content_type(img_url)
                        key = _b2_key(cfg.prefix, source, listing_id, idx, sha1, ct)
                        out.append({"b2_key": key, "file_name": os.path.basename(key), "content_type": ct, "source_url": img_url})
                        continue

                    data, ct = _download_image(img_url)
                    sha1 = _sha1_bytes(data)
                    key = _b2_key(cfg.prefix, source, listing_id, idx, sha1, ct)

                    # Upload (small files: upload_bytes)
                    bucket.upload_bytes(
                        data,
                        file_name=key,
                        content_type=ct,
                        file_info={
                            "source": source,
                            "source_listing_id": listing_id,
                            "listing_url": listing_url,
                            "source_url": img_url,
                            "sha1": sha1,
                        },
                    )
                    out.append({"b2_key": key, "file_name": os.path.basename(key), "content_type": ct, "source_url": img_url})

                    time.sleep(delay)

                _row_set(ws, r, headers, "b2_images_json", json.dumps(out, ensure_ascii=False))
                _row_set(ws, r, headers, "b2_status", "OK")
                _row_set(ws, r, headers, "b2_error", "")
                _row_set(ws, r, headers, "updated_at", time.strftime("%Y-%m-%dT%H:%M:%S"))

            except Exception as e:
                _row_set(ws, r, headers, "b2_status", "ERROR")
                _row_set(ws, r, headers, "b2_error", str(e))
                _row_set(ws, r, headers, "updated_at", time.strftime("%Y-%m-%dT%H:%M:%S"))

            processed += 1

        wb.save(xlsx_path)

    if not found_any:
        raise RuntimeError(f"No detail sheets found. Looked for: {sheets}. Workbook sheets: {wb.sheetnames}")

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True, help="Path to workbook (same one your scrapers write).")
    p.add_argument("--sheets", default=",".join(DEFAULT_DETAIL_SHEETS),
                   help="Comma-separated sheet names to ingest. Default: SS_Detail_Enrichment,JBG_Detail_Enrichment")
    p.add_argument("--limit", type=int, default=0, help="Max rows to process per run (0 = no limit).")
    p.add_argument("--delay", type=float, default=0.5, help="Delay between image uploads (seconds).")
    p.add_argument("--resume", action="store_true", help="Skip rows that already have b2_images_json.")
    p.add_argument("--dry-run", action="store_true", help="Do not download/upload; just compute expected B2 keys.")
    args = p.parse_args()

    sheets = [s.strip() for s in args.sheets.split(",") if s.strip()]
    ingest_images(args.xlsx, sheets, args.limit, args.delay, args.resume, args.dry_run)

if __name__ == "__main__":
    main()
