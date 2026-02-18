#!/usr/bin/env python3
"""
GloveIQ Library Import

Reads scraper workbook XLSX and writes deterministic exports:
- listings.normalized.jsonl
- listings.raw.jsonl
- media_manifest.jsonl
- import_report.json

Design goals:
- stable listing primary key: source + source_listing_id
- deterministic ordering and keys
- resumable via input fingerprint checkpoint
- no secret handling; optional B2 prefix only from env/CLI
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import mimetypes
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse

from openpyxl import load_workbook


REQUIRED_SHEETS: Dict[str, List[str]] = {
    "Catalog": [
        "listing_id",
        "product_url",
        "title",
        "price",
        "currency",
        "condition",
        "brand",
        "model",
        "images_json",
        "normalized_json",
    ],
    "JBG_Full_Catalog": [
        "product_id",
        "product_url",
        "source",
        "catalog_title",
        "catalog_price",
        "thumb_url",
        "catalog_scraped_at",
    ],
    "JBG_Detail_Enrichment": [
        "product_id",
        "product_url",
        "detail_scraped_at",
        "detail_status",
        "detail_error",
        "title",
        "price",
        "model_code",
        "glove_profile_json",
        "description_snippet",
        "images_json",
        "spec_json",
    ],
}

KNOWN_BRANDS = [
    "Wilson",
    "Rawlings",
    "Mizuno",
    "Easton",
    "Marucci",
    "Franklin",
    "Louisville Slugger",
    "Nike",
    "Nokona",
    "SSK",
    "Adidas",
    "All Star",
    "Akadema",
    "44 Pro",
]

REQUIRED_SPEC_FIELDS = [
    "Item #",
    "Back",
    "Color",
    "Fit",
    "Leather",
    "Level",
    "Lining",
    "Padding",
    "Pattern",
    "Series",
    "Shell",
    "Size",
    "Special Feature",
    "Sport",
    "Throwing Hand",
    "Usage",
    "Used by",
    "Web",
    "Wrist",
    "Age Group",
    "Description",
]


@dataclass
class ValidationResult:
    ok: bool
    errors: List[str]
    warnings: List[str]


def _now_iso() -> str:
    return dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _clean(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _safe_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        s = _clean(v)
        if not s:
            return None
        m = re.search(r"([\d,]+(?:\.\d+)?)", s)
        if not m:
            return None
        try:
            return float(m.group(1).replace(",", ""))
        except Exception:
            return None


def _safe_json(v: Any, fallback: Any) -> Any:
    if v is None:
        return fallback
    if isinstance(v, (dict, list)):
        return v
    s = _clean(v)
    if not s:
        return fallback
    try:
        return json.loads(s)
    except Exception:
        return fallback


def _header_map(ws) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        k = _clean(v)
        if k:
            out[k] = c
    return out


def validate_workbook(xlsx_path: str) -> ValidationResult:
    errors: List[str] = []
    warnings: List[str] = []

    wb = load_workbook(xlsx_path, read_only=False, data_only=True)

    for sheet, required_cols in REQUIRED_SHEETS.items():
        if sheet not in wb.sheetnames:
            errors.append(f"Missing required sheet: {sheet}")
            continue
        ws = wb[sheet]
        headers = _header_map(ws)
        for col in required_cols:
            if col not in headers:
                errors.append(f"Sheet {sheet} missing required column: {col}")

    if errors:
        return ValidationResult(ok=False, errors=errors, warnings=warnings)

    # Row-level checks
    ws_catalog = wb["Catalog"]
    hc = _header_map(ws_catalog)
    for r in range(2, ws_catalog.max_row + 1):
        listing_id = _clean(ws_catalog.cell(r, hc["listing_id"]).value)
        url = _clean(ws_catalog.cell(r, hc["product_url"]).value)
        if any(_clean(ws_catalog.cell(r, c).value) for c in hc.values()):
            if not listing_id:
                errors.append(f"Catalog row {r} missing listing_id")
            if not url:
                errors.append(f"Catalog row {r} missing product_url")

    ws_jcat = wb["JBG_Full_Catalog"]
    hjc = _header_map(ws_jcat)
    for r in range(2, ws_jcat.max_row + 1):
        pid = _clean(ws_jcat.cell(r, hjc["product_id"]).value)
        url = _clean(ws_jcat.cell(r, hjc["product_url"]).value)
        if any(_clean(ws_jcat.cell(r, c).value) for c in hjc.values()):
            if not pid:
                errors.append(f"JBG_Full_Catalog row {r} missing product_id")
            if not url:
                errors.append(f"JBG_Full_Catalog row {r} missing product_url")

    ws_jdet = wb["JBG_Detail_Enrichment"]
    hjd = _header_map(ws_jdet)
    for r in range(2, ws_jdet.max_row + 1):
        pid = _clean(ws_jdet.cell(r, hjd["product_id"]).value)
        url = _clean(ws_jdet.cell(r, hjd["product_url"]).value)
        if any(_clean(ws_jdet.cell(r, c).value) for c in hjd.values()):
            if not pid:
                errors.append(f"JBG_Detail_Enrichment row {r} missing product_id")
            if not url:
                errors.append(f"JBG_Detail_Enrichment row {r} missing product_url")

    return ValidationResult(ok=(len(errors) == 0), errors=errors, warnings=warnings)


def _norm_throw(v: Optional[str]) -> Optional[str]:
    s = (_clean(v) or "").lower()
    if not s:
        return None
    if "rht" in s or "right" in s:
        return "RHT"
    if "lht" in s or "left" in s:
        return "LHT"
    return "UNK"


def _norm_position(v: Optional[str]) -> Optional[str]:
    s = (_clean(v) or "").lower()
    if not s:
        return None
    if "outfield" in s or s == "of":
        return "OF"
    if "infield" in s or s == "if":
        return "IF"
    if "pitch" in s:
        return "P"
    if "catch" in s:
        return "C"
    if "1b" in s or "first" in s:
        return "1B"
    if "middle" in s or "ss" in s or "2b" in s:
        return "MI"
    if "utility" in s:
        return "Utility"
    return _clean(v)


def _extract_size_in(text: Optional[str]) -> Optional[float]:
    s = _clean(text)
    if not s:
        return None
    m = re.search(r"(\d{1,2}\.\d{1,2})", s)
    if m:
        return _safe_float(m.group(1))
    m = re.search(r"(\d{1,2})\s*[\-\s]\s*(\d)\s*/\s*(\d)", s)
    if m:
        return float(m.group(1)) + (float(m.group(2)) / float(m.group(3)))
    m = re.search(r"(\d{1,2}(?:\.\d+)?)\s*\"", s)
    if m:
        return _safe_float(m.group(1))
    return None


def _infer_brand(title: Optional[str], explicit: Optional[str]) -> Optional[str]:
    if _clean(explicit):
        return _clean(explicit)
    t = (_clean(title) or "").lower()
    for brand in KNOWN_BRANDS:
        if brand.lower() in t:
            return brand
    return None


def _infer_model(title: Optional[str], explicit: Optional[str]) -> Optional[str]:
    if _clean(explicit):
        return _clean(explicit)
    t = _clean(title)
    if not t:
        return None
    m = re.search(r"\[Model:\s*([^\]]+)\]", t, re.I)
    if m:
        return _clean(m.group(1))
    m = re.search(r"\(([A-Z0-9\-]{5,})\)\s*$", t)
    if m:
        return _clean(m.group(1))
    return None


def _infer_sport(title: Optional[str], profile: Dict[str, Any]) -> str:
    title_l = (_clean(title) or "").lower()
    text = " ".join(str(v) for v in profile.values()).lower()
    merged = f"{title_l} {text}"
    if "softball" in merged or "fastpitch" in merged:
        return "softball"
    return "baseball"


def _norm_images(v: Any) -> List[str]:
    arr = _safe_json(v, [])
    out: List[str] = []
    if not isinstance(arr, list):
        return out
    seen = set()
    for item in arr:
        u = _clean(item)
        if not u or not (u.startswith("http://") or u.startswith("https://")):
            continue
        if u in seen:
            continue
        seen.add(u)
        out.append(u)
    return out


def _slug(s: Optional[str]) -> str:
    raw = (_clean(s) or "unknown").lower()
    return re.sub(r"[^a-z0-9]+", "-", raw).strip("-") or "unknown"


def _record_type_from_listing(*, source: str, condition: Optional[str], model_code: Optional[str], title: Optional[str]) -> str:
    text = " ".join([source or "", condition or "", model_code or "", title or ""]).lower()
    artifact_markers = ["custom", "one of one", "1/1", "game used", "player issued", "modified", "re-lace", "relace"]
    if any(marker in text for marker in artifact_markers):
        return "artifact"
    if model_code and _clean(model_code) and _clean(model_code) != "Unknown":
        return "variant"
    if source.upper() in {"JBG", "JUSTBALLGLOVES"}:
        return "variant"
    return "artifact"


def _build_spec_map(
    title: Optional[str],
    description: Optional[str],
    model_code: Optional[str],
    size_in: Optional[float],
    throw_hand: Optional[str],
    sport: Optional[str],
    web_type: Optional[str],
    glove_profile: Optional[Dict[str, Any]] = None,
    spec_json: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[str, Optional[str]], Dict[str, float]]:
    by_lower: Dict[str, str] = {}
    for container in (glove_profile or {}, spec_json or {}):
        for k, v in container.items():
            kk = _clean(k)
            vv = _clean(v)
            if kk and vv:
                by_lower[kk.lower()] = vv

    def from_any(*keys: str) -> Optional[str]:
        for key in keys:
            hit = by_lower.get(key.lower())
            if hit:
                return hit
        return None

    out: Dict[str, Optional[str]] = {
        "Item #": _clean(model_code),
        "Back": from_any("Back"),
        "Color": from_any("Color"),
        "Fit": from_any("Fit"),
        "Leather": from_any("Leather"),
        "Level": from_any("Level"),
        "Lining": from_any("Lining"),
        "Padding": from_any("Padding"),
        "Pattern": from_any("Pattern"),
        "Series": from_any("Series"),
        "Shell": from_any("Shell"),
        "Size": from_any("Size") or (f"{size_in:.2f}" if isinstance(size_in, (int, float)) else None),
        "Special Feature": from_any("Special Feature"),
        "Sport": from_any("Sport") or _clean(sport),
        "Throwing Hand": from_any("Throwing Hand") or _clean(throw_hand),
        "Usage": from_any("Usage"),
        "Used by": from_any("Used by"),
        "Web": from_any("Web") or _clean(web_type),
        "Wrist": from_any("Wrist"),
        "Age Group": from_any("Age Group"),
        "Description": _clean(description) or _clean(title),
    }

    confidence: Dict[str, float] = {}
    for field in REQUIRED_SPEC_FIELDS:
        value = out.get(field)
        confidence[field] = 0.92 if value else 0.0
    return out, confidence


def _guess_ext_and_ct(url: str) -> Tuple[str, str]:
    no_query = url.split("?", 1)[0]
    ct, _ = mimetypes.guess_type(no_query)
    ct = ct or "image/jpeg"
    ext = mimetypes.guess_extension(ct) or Path(urlparse(no_query).path).suffix or ".jpg"
    if not ext.startswith("."):
        ext = f".{ext}"
    return ext, ct


def _image_target_key(prefix: str, source: str, listing_id: str, index_1: int, image_url: str) -> Tuple[str, str]:
    sha1 = hashlib.sha1(image_url.encode("utf-8")).hexdigest()
    ext, ct = _guess_ext_and_ct(image_url)
    key = f"{prefix}/{source}/{listing_id}/{index_1:02d}_{sha1[:10]}{ext}"
    return key, ct


def _iter_sheet_rows(ws, headers: Dict[str, int]) -> Iterable[Tuple[int, Dict[str, Any]]]:
    for r in range(2, ws.max_row + 1):
        row = {k: ws.cell(r, c).value for k, c in headers.items()}
        if not any(_clean(v) for v in row.values()):
            continue
        yield r, row


def _stable_key(source: str, listing_id: str) -> str:
    return f"{source}:{listing_id}"


def _build_jbg_catalog_index(ws) -> Dict[str, Dict[str, Any]]:
    h = _header_map(ws)
    out: Dict[str, Dict[str, Any]] = {}
    for r, row in _iter_sheet_rows(ws, h):
        pid = _clean(row.get("product_id"))
        if not pid:
            continue
        out[pid] = {
            "sheet": ws.title,
            "row": r,
            "product_url": _clean(row.get("product_url")),
            "source": _clean(row.get("source")) or "JBG",
            "catalog_title": _clean(row.get("catalog_title")),
            "catalog_price": _safe_float(row.get("catalog_price")),
            "thumb_url": _clean(row.get("thumb_url")),
            "catalog_scraped_at": _clean(row.get("catalog_scraped_at")),
            "raw": row,
        }
    return out


def build_exports(xlsx_path: str, b2_prefix: str) -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, read_only=False, data_only=True)

    listings: List[Dict[str, Any]] = []
    raw_rows: List[Dict[str, Any]] = []
    media_rows: List[Dict[str, Any]] = []
    errors: List[str] = []
    defaults_applied: Dict[str, int] = {}

    by_source = {"SS": 0, "JBG": 0}
    rows_scanned = {"Catalog": 0, "JBG_Full_Catalog": 0, "JBG_Detail_Enrichment": 0}

    # SS / Catalog
    ws_ss = wb["Catalog"]
    hs = _header_map(ws_ss)
    for r, row in _iter_sheet_rows(ws_ss, hs):
        rows_scanned["Catalog"] += 1
        listing_id = _clean(row.get("listing_id"))
        url = _clean(row.get("product_url"))
        if not listing_id or not url:
            errors.append(f"Catalog row {r} missing listing_id/product_url")
            continue

        norm = _safe_json(row.get("normalized_json"), {})
        norm_raw = norm.get("raw", {}) if isinstance(norm, dict) else {}
        norm_obj = norm.get("norm", {}) if isinstance(norm, dict) else {}

        title = _clean(row.get("title")) or _clean(norm_obj.get("title"))
        brand = _infer_brand(title, _clean(row.get("brand")))
        model = _infer_model(title, _clean(row.get("model")))
        size_in = _safe_float(norm_obj.get("size_in")) if isinstance(norm_obj, dict) else None
        throw_hand = _norm_throw(_clean(norm_obj.get("throw_hand")) if isinstance(norm_obj, dict) else None)
        position = _norm_position(_clean(norm_obj.get("position")) if isinstance(norm_obj, dict) else None)
        web_type = _clean(norm_obj.get("web")) if isinstance(norm_obj, dict) else None
        description = _clean(norm_obj.get("description")) if isinstance(norm_obj, dict) else None
        record_type = _record_type_from_listing(
            source="SS",
            condition=_clean(row.get("condition")),
            model_code=model,
            title=title,
        )
        canonical_name = " ".join([x for x in [brand, model, f"{size_in:.2f}" if isinstance(size_in, (int, float)) else None] if x]).strip() or title or "Unknown"
        glove_id = (
            f"variant:{_slug(brand)}:{_slug(model)}:{_slug(str(size_in) if size_in is not None else 'na')}:{_slug(throw_hand or 'unk')}"
            if record_type == "variant"
            else f"artifact:SS:{listing_id}"
        )
        specs_raw, specs_conf = _build_spec_map(
            title=title,
            description=description,
            model_code=model,
            size_in=size_in,
            throw_hand=throw_hand,
            sport=_infer_sport(title, {}),
            web_type=web_type,
            glove_profile={},
            spec_json=norm_raw if isinstance(norm_raw, dict) else {},
        )

        if not brand:
            defaults_applied["brand_unknown"] = defaults_applied.get("brand_unknown", 0) + 1
        if not model:
            defaults_applied["model_unknown"] = defaults_applied.get("model_unknown", 0) + 1

        images = _norm_images(row.get("images_json"))
        listing = {
            "listing_pk": _stable_key("SS", listing_id),
            "glove_id": glove_id,
            "record_type": record_type,
            "source": "SS",
            "source_listing_id": listing_id,
            "url": url,
            "title": title,
            "canonical_name": canonical_name,
            "brand": brand or "Unknown",
            "model": model or "Unknown",
            "model_code": model or "Unknown",
            "size_in": size_in,
            "hand": throw_hand or "UNK",
            "throw_hand": throw_hand or "UNK",
            "player_position": position or "Unknown",
            "position": position or "Unknown",
            "web_type": web_type or "Unknown",
            "sport": _infer_sport(title, {}),
            "condition": _clean(row.get("condition")) or "Unknown",
            "price": _safe_float(row.get("price")),
            "currency": _clean(row.get("currency")) or "USD",
            "created_at": None,
            "seen_at": None,
            "item_number": model or None,
            "pattern": _clean(norm_obj.get("pattern")) if isinstance(norm_obj, dict) else None,
            "series": _clean(norm_obj.get("series")) if isinstance(norm_obj, dict) else None,
            "level": _clean(norm_obj.get("level")) if isinstance(norm_obj, dict) else None,
            "age_group": _clean(norm_obj.get("age_group")) if isinstance(norm_obj, dict) else None,
            "market_origin": None,
            "raw_specs": norm_raw if isinstance(norm_raw, dict) else {},
            "spec_fields_raw": specs_raw,
            "normalized_specs": {k: v for k, v in specs_raw.items() if v},
            "normalized_confidence": specs_conf,
            "raw_html": None,
            "raw_text": title,
            "images": images,
        }
        listings.append(listing)
        by_source["SS"] += 1

        raw_rows.append(
            {
                "listing_pk": listing["listing_pk"],
                "source": "SS",
                "source_sheet": "Catalog",
                "source_row": r,
                "source_columns": row,
                "parsed_normalized_json": norm,
                "raw_html": None,
                "raw_text": title,
            }
        )

    # JBG index + details
    ws_jcat = wb["JBG_Full_Catalog"]
    for _ in _iter_sheet_rows(ws_jcat, _header_map(ws_jcat)):
        rows_scanned["JBG_Full_Catalog"] += 1
    jbg_catalog_idx = _build_jbg_catalog_index(ws_jcat)

    ws_jdet = wb["JBG_Detail_Enrichment"]
    hd = _header_map(ws_jdet)
    for r, row in _iter_sheet_rows(ws_jdet, hd):
        rows_scanned["JBG_Detail_Enrichment"] += 1
        pid = _clean(row.get("product_id"))
        detail_url = _clean(row.get("product_url"))
        if not pid:
            errors.append(f"JBG_Detail_Enrichment row {r} missing product_id")
            continue

        cat = jbg_catalog_idx.get(pid, {})
        source = _clean(cat.get("source")) or "JBG"
        url = detail_url or _clean(cat.get("product_url"))
        if not url:
            errors.append(f"JBG listing {pid} missing URL in catalog+detail")
            continue

        glove_profile = _safe_json(row.get("glove_profile_json"), {})
        spec_json = _safe_json(row.get("spec_json"), {})

        title = _clean(row.get("title")) or _clean(cat.get("catalog_title"))
        price = _safe_float(row.get("price"))
        if price is None:
            price = _safe_float(cat.get("catalog_price"))
            if price is not None:
                defaults_applied["price_from_catalog"] = defaults_applied.get("price_from_catalog", 0) + 1

        brand = _infer_brand(title, None)
        model_code = _clean(row.get("model_code")) or _infer_model(title, None)
        model = model_code or _infer_model(title, None)

        size_text = None
        throw_text = None
        pos_text = None
        web_text = None

        if isinstance(glove_profile, dict):
            for k, v in glove_profile.items():
                lk = (_clean(k) or "").lower()
                if "size" in lk and not size_text:
                    size_text = _clean(v)
                if ("throw" in lk or "hand" in lk) and not throw_text:
                    throw_text = _clean(v)
                if "position" in lk and not pos_text:
                    pos_text = _clean(v)
                if "web" in lk and not web_text:
                    web_text = _clean(v)

        size_in = _extract_size_in(size_text) or _extract_size_in(title)
        throw_hand = _norm_throw(throw_text)
        position = _norm_position(pos_text)

        if not brand:
            defaults_applied["brand_unknown"] = defaults_applied.get("brand_unknown", 0) + 1
        if not model:
            defaults_applied["model_unknown"] = defaults_applied.get("model_unknown", 0) + 1

        images = _norm_images(row.get("images_json"))
        description = _clean(row.get("description_snippet"))
        sport = _infer_sport(title, glove_profile if isinstance(glove_profile, dict) else {})
        condition = "New" if source == "JBG" else "Unknown"
        record_type = _record_type_from_listing(
            source=source,
            condition=condition,
            model_code=model_code,
            title=title,
        )
        canonical_name = " ".join([x for x in [brand, model_code or model, f"{size_in:.2f}" if isinstance(size_in, (int, float)) else None] if x]).strip() or title or "Unknown"
        glove_id = (
            f"variant:{_slug(brand)}:{_slug(model_code or model)}:{_slug(str(size_in) if size_in is not None else 'na')}:{_slug(throw_hand or 'unk')}"
            if record_type == "variant"
            else f"artifact:{source}:{pid}"
        )
        specs_raw, specs_conf = _build_spec_map(
            title=title,
            description=description,
            model_code=model_code or model,
            size_in=size_in,
            throw_hand=throw_hand,
            sport=sport,
            web_type=web_text,
            glove_profile=glove_profile if isinstance(glove_profile, dict) else {},
            spec_json=spec_json if isinstance(spec_json, dict) else {},
        )

        listing = {
            "listing_pk": _stable_key(source, pid),
            "glove_id": glove_id,
            "record_type": record_type,
            "source": source,
            "source_listing_id": pid,
            "url": url,
            "title": title,
            "canonical_name": canonical_name,
            "brand": brand or "Unknown",
            "model": model or "Unknown",
            "model_code": model_code or "Unknown",
            "size_in": size_in,
            "hand": throw_hand or "UNK",
            "throw_hand": throw_hand or "UNK",
            "player_position": position or "Unknown",
            "position": position or "Unknown",
            "web_type": web_text or "Unknown",
            "sport": sport,
            "condition": condition,
            "price": price,
            "currency": "USD",
            "created_at": _clean(cat.get("catalog_scraped_at")),
            "seen_at": _clean(row.get("detail_scraped_at")),
            "item_number": model_code or None,
            "pattern": _clean(glove_profile.get("pattern")) if isinstance(glove_profile, dict) else None,
            "series": _clean(glove_profile.get("series")) if isinstance(glove_profile, dict) else None,
            "level": _clean(glove_profile.get("level")) if isinstance(glove_profile, dict) else None,
            "age_group": _clean(glove_profile.get("age_group")) if isinstance(glove_profile, dict) else None,
            "market_origin": _clean(glove_profile.get("country")) if isinstance(glove_profile, dict) else None,
            "raw_specs": {
                "glove_profile": glove_profile if isinstance(glove_profile, dict) else {},
                "spec_json": spec_json if isinstance(spec_json, dict) else {},
            },
            "spec_fields_raw": specs_raw,
            "normalized_specs": {k: v for k, v in specs_raw.items() if v},
            "normalized_confidence": specs_conf,
            "raw_html": None,
            "raw_text": _clean(row.get("description_snippet")) or title,
            "images": images,
        }
        listings.append(listing)
        by_source["JBG"] += 1

        raw_rows.append(
            {
                "listing_pk": listing["listing_pk"],
                "source": source,
                "source_sheet": "JBG_Detail_Enrichment",
                "source_row": r,
                "source_columns": row,
                "catalog_columns": cat.get("raw", {}),
                "parsed_glove_profile_json": glove_profile,
                "parsed_spec_json": spec_json,
                "raw_html": None,
                "raw_text": _clean(row.get("description_snippet")) or title,
            }
        )

    # Deduplicate by listing key (last writer wins), then sorted deterministic output
    dedup: Dict[str, Dict[str, Any]] = {}
    for l in listings:
        dedup[l["listing_pk"]] = l
    listings_sorted = [dedup[k] for k in sorted(dedup.keys())]

    # Keep raw rows only for retained listing keys (last raw for each key)
    raw_dedup: Dict[str, Dict[str, Any]] = {}
    for rr in raw_rows:
        raw_dedup[rr["listing_pk"]] = rr
    raw_sorted = [raw_dedup[k] for k in sorted(raw_dedup.keys())]

    # Media manifest (one row per listing with ordered image list + deterministic b2 key mapping)
    prefix = (b2_prefix or "gloveiq").strip().strip("/")
    image_total = 0
    for l in listings_sorted:
        mappings = []
        for idx, img_url in enumerate(l.get("images") or [], start=1):
            target_key, content_type = _image_target_key(prefix, l["source"], l["source_listing_id"], idx, img_url)
            mappings.append(
                {
                    "image_index": idx,
                    "source_url": img_url,
                    "target_storage_key": target_key,
                    "content_type": content_type,
                    "mapping_key": f"{l['source']}:{l['source_listing_id']}:{idx}",
                }
            )
            image_total += 1

        media_rows.append(
            {
                "listing_pk": l["listing_pk"],
                "source": l["source"],
                "source_listing_id": l["source_listing_id"],
                "ordered_image_urls": l.get("images") or [],
                "image_mappings": mappings,
            }
        )

    by_source_final: Dict[str, int] = {}
    by_record_type: Dict[str, int] = {}
    for listing in listings_sorted:
        src = listing.get("source") or "Unknown"
        by_source_final[src] = by_source_final.get(src, 0) + 1
        rtype = listing.get("record_type") or "artifact"
        by_record_type[rtype] = by_record_type.get(rtype, 0) + 1

    report = {
        "generated_at": _now_iso(),
        "input_xlsx": os.path.abspath(xlsx_path),
        "rows_scanned": rows_scanned,
        "listings_total": len(listings_sorted),
        "listings_by_source": by_source_final,
        "listings_by_record_type": by_record_type,
        "media_manifest_rows": len(media_rows),
        "media_manifest_images_total": image_total,
        "errors_count": len(errors),
        "errors": errors[:500],
        "warnings_count": 0,
        "warnings": [],
        "defaults_applied": defaults_applied,
    }

    return {
        "listings": listings_sorted,
        "raw_rows": raw_sorted,
        "media_manifest": media_rows,
        "report": report,
    }


def _write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")
    tmp.replace(path)


def _file_fingerprint(path: str) -> Dict[str, Any]:
    st = os.stat(path)
    return {
        "path": os.path.abspath(path),
        "size": st.st_size,
        "mtime": int(st.st_mtime),
    }


def run_import(
    xlsx: str,
    out_dir: str,
    b2_prefix: str,
    emit_raw: bool,
    resume: bool,
    force: bool,
) -> int:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    checkpoint_path = out / ".library_import.checkpoint.json"
    normalized_path = out / "listings.normalized.jsonl"
    raw_path = out / "listings.raw.jsonl"
    media_path = out / "media_manifest.jsonl"
    report_path = out / "import_report.json"

    fingerprint = _file_fingerprint(xlsx)
    if resume and not force and checkpoint_path.exists() and normalized_path.exists() and media_path.exists() and report_path.exists():
        old = _safe_json(checkpoint_path.read_text(encoding="utf-8"), {})
        if isinstance(old, dict) and old.get("input_fingerprint") == fingerprint:
            print("[library_import] unchanged input fingerprint, skipping export (use --force to regenerate)")
            return 0

    validation = validate_workbook(xlsx)
    if not validation.ok:
        print("[library_import] validation failed:")
        for e in validation.errors:
            print(f"  - {e}")
        return 2

    exports = build_exports(xlsx, b2_prefix=b2_prefix)
    _write_jsonl(normalized_path, exports["listings"])
    if emit_raw:
        _write_jsonl(raw_path, exports["raw_rows"])
    _write_jsonl(media_path, exports["media_manifest"])

    report_payload = exports["report"]
    report_payload["output_files"] = {
        "normalized": str(normalized_path),
        "raw": str(raw_path) if emit_raw else None,
        "manifest": str(media_path),
    }
    report_path.write_text(json.dumps(report_payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n", encoding="utf-8")

    checkpoint_path.write_text(
        json.dumps(
            {
                "generated_at": _now_iso(),
                "input_fingerprint": fingerprint,
                "counts": {
                    "listings": len(exports["listings"]),
                    "media_rows": len(exports["media_manifest"]),
                },
            },
            indent=2,
            ensure_ascii=False,
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )

    print(f"[library_import] wrote {normalized_path}")
    if emit_raw:
        print(f"[library_import] wrote {raw_path}")
    print(f"[library_import] wrote {media_path}")
    print(f"[library_import] wrote {report_path}")
    return 0


def main() -> None:
    p = argparse.ArgumentParser(description="Generate GloveIQ library import exports from scraper workbook XLSX")
    p.add_argument("--xlsx", required=True, help="Workbook path")
    p.add_argument("--out-dir", default="data_exports", help="Output folder")
    p.add_argument("--b2-prefix", default=os.getenv("B2_PREFIX", "gloveiq"), help="B2 key prefix used in manifest")
    p.add_argument("--no-raw", action="store_true", help="Disable listings.raw.jsonl output")
    p.add_argument("--no-resume", action="store_true", help="Always regenerate outputs")
    p.add_argument("--force", action="store_true", help="Force regeneration even if fingerprint unchanged")
    args = p.parse_args()

    code = run_import(
        xlsx=args.xlsx,
        out_dir=args.out_dir,
        b2_prefix=args.b2_prefix,
        emit_raw=not args.no_raw,
        resume=not args.no_resume,
        force=args.force,
    )
    raise SystemExit(code)


if __name__ == "__main__":
    main()
