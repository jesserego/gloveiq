#!/usr/bin/env python3
"""
JustBallGloves (JustGloves) scraper -> GloveIQ template XLSX

Writes to:
- Sheet: JBG_Full_Catalog
  - one row per product discovered from catalog pages
- Sheet: JBG_Detail_Enrichment
  - one row per product detail scrape (specs + images_json)

Usage examples:

# 1) Catalog only (discover products) — fast
python jbg_master_scraper.py --xlsx "../jbg/GloveIQ_Library_Master_Template_GOOGLE_NATIVE_FULLCAT_READY.xlsx" --max-pages 25 --max-details 0 --delay 1.5

# 2) Details only (enrich existing catalog rows) — slower
python jbg_master_scraper.py --xlsx "../jbg/GloveIQ_Library_Master_Template_GOOGLE_NATIVE_FULLCAT_READY.xlsx" --max-pages 0 --max-details 500 --delay 1.5 --resume

# 3) Full run (catalog + details)
python jbg_master_scraper.py --xlsx "../jbg/GloveIQ_Library_Master_Template_GOOGLE_NATIVE_FULLCAT_READY.xlsx" --max-pages 25 --max-details 2000 --delay 1.5
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import re
import time
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from urllib.parse import urljoin, urlparse, urlunparse, parse_qs, urlencode

import requests
from bs4 import BeautifulSoup
import openpyxl


DEFAULT_START_URL = "https://www.justballgloves.com/products/glove%20type~baseball,female%20fastpitch,slow%20pitch%20softball,softball,tee%20ball,youth/"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36"

PRODUCT_RE = re.compile(r"/(?:product|products)/(?:[^/]+/)*?(?P<pid>\d+)(?:/|$)", re.I)


def now_iso() -> str:
    return _dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def abs_url(base: str, href: str) -> str:
    return urljoin(base, href)


def set_query_param(url: str, key: str, value: str) -> str:
    u = urlparse(url)
    q = parse_qs(u.query)
    q[key] = [value]
    new_q = urlencode(q, doseq=True)
    return urlunparse((u.scheme, u.netloc, u.path, u.params, new_q, u.fragment))


def safe_json(obj) -> str:
    return json.dumps(obj, ensure_ascii=False, separators=(",", ":"))


def get_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(
        {
            "User-Agent": UA,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        }
    )
    return s


def load_wb(xlsx_path: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(xlsx_path)


def header_map(ws) -> Dict[str, int]:
    """Return {header_name: col_index} for row 1."""
    m = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            m[name] = col
    return m


def first_empty_row(ws, key_col: int) -> int:
    """Find first empty row by checking a primary key column (e.g., product_id)."""
    r = 2
    while True:
        if ws.cell(row=r, column=key_col).value in (None, ""):
            return r
        r += 1


def build_existing_index(ws, key_header: str) -> Dict[str, int]:
    hm = header_map(ws)
    if key_header not in hm:
        raise RuntimeError(f"Sheet {ws.title} missing required header: {key_header}")
    key_col = hm[key_header]
    idx = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is None or str(v).strip() == "":
            continue
        idx[str(v).strip()] = r
    return idx


def parse_money(text: str) -> Optional[float]:
    if not text:
        return None
    m = re.search(r"\$([\d,]+(?:\.\d+)?)", text)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", ""))
    except Exception:
        return None


def extract_catalog_products(html: str, base_url: str) -> List[Dict]:
    soup = BeautifulSoup(html, "lxml")
    found = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        m = PRODUCT_RE.search(href)
        if not m:
            continue
        pid = m.group("pid")
        url = abs_url(base_url, href)
        # try to pick a reasonable title/price from surrounding card
        card = a
        for _ in range(4):
            if getattr(card, "name", None) in ("article", "li", "div"):
                break
            card = card.parent
            if card is None:
                break

        title = None
        price = None
        thumb = None

        # Title: look for h2/h3 or data-name
        if card:
            h = card.find(["h2", "h3"])
            if h and h.get_text(strip=True):
                title = h.get_text(" ", strip=True)

            # Price: any element with $ in it inside the card
            t = card.get_text(" ", strip=True)
            price = parse_money(t)

            img = card.find("img")
            if img:
                thumb = img.get("data-src") or img.get("src")

        found.append(
            {
                "product_id": pid,
                "product_url": url,
                "title_catalog": title,
                "price_catalog": price,
                "thumb_url": thumb,
            }
        )

    # de-dupe by product_id while preserving order
    dedup = {}
    out = []
    for item in found:
        if item["product_id"] in dedup:
            continue
        dedup[item["product_id"]] = True
        out.append(item)
    return out


def parse_glove_profile(soup: BeautifulSoup) -> Dict[str, str]:
    """
    Extract 'Glove Profile' key/value pairs.
    Site often uses label + value (e.g., 'Positions:' then chips/values).
    """
    profile = {}

    # Find section heading containing 'Glove Profile'
    h = soup.find(lambda tag: tag.name in ("h2", "h3") and "glove profile" in tag.get_text(" ", strip=True).lower())
    if not h:
        return profile

    # Section container: climb to a reasonable wrapper
    section = h
    for _ in range(6):
        if section.parent is None:
            break
        section = section.parent
        txt = section.get_text(" ", strip=True).lower()
        if "glove profile" in txt and ("positions" in txt or "age" in txt or "stiffness" in txt):
            break

    # Heuristic: label elements frequently end with ':'
    text = section.get_text("\n", strip=True)
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]

    # Build a loose key/value parser:
    # If line ends with ':' treat as key; next non-key line is value (possibly multiple tokens).
    i = 0
    while i < len(lines):
        ln = lines[i]
        if ln.endswith(":") and len(ln) < 40:
            key = ln[:-1].strip()
            # accumulate subsequent non-key lines until next key-like line
            vals = []
            j = i + 1
            while j < len(lines):
                nxt = lines[j]
                if nxt.endswith(":") and len(nxt) < 40:
                    break
                # avoid capturing the heading itself
                if nxt.lower() != "glove profile":
                    vals.append(nxt)
                j += 1
            if vals:
                profile[key] = " ".join(vals).strip()
            i = j
        else:
            i += 1

    return profile


def parse_images(soup: BeautifulSoup, base_url: str) -> List[str]:
    urls = []
    for img in soup.find_all("img"):
        src = img.get("data-src") or img.get("data-lazy-src") or img.get("src")
        if not src:
            continue
        if "justballgloves" not in src and "cdn" not in src and not src.startswith("http"):
            continue
        full = abs_url(base_url, src)
        urls.append(full)

    # de-dupe; keep order
    seen = set()
    out = []
    for u in urls:
        if u in seen:
            continue
        seen.add(u)
        out.append(u)
    return out


def parse_detail_page(html: str, url: str) -> Dict:
    soup = BeautifulSoup(html, "lxml")
    title = None
    h1 = soup.find("h1")
    if h1 and h1.get_text(strip=True):
        title = h1.get_text(" ", strip=True)

    # price: try common patterns first
    price = None
    # Many commerce templates store prices in meta
    meta_price = soup.find("meta", {"property": "product:price:amount"}) or soup.find("meta", {"itemprop": "price"})
    if meta_price and meta_price.get("content"):
        try:
            price = float(meta_price["content"])
        except Exception:
            price = None
    if price is None:
        # fallback: any visible $ near the top
        price = parse_money(soup.get_text(" ", strip=True)[:5000])

    glove_profile = parse_glove_profile(soup)
    images = parse_images(soup, url)

    # Try to extract a "model_code" from parentheses at end of title, e.g. "(WBW100396115)"
    model_code = None
    if title:
        m = re.search(r"\(([A-Z0-9\-]{6,})\)\s*$", title)
        if m:
            model_code = m.group(1)

    # Try to capture bullet benefits or description snippets (lightweight)
    desc = None
    # common: element with id/section containing "Glove Benefits"
    hb = soup.find(lambda tag: tag.name in ("h2", "h3") and "glove benefits" in tag.get_text(" ", strip=True).lower())
    if hb:
        sec = hb
        for _ in range(6):
            if sec.parent is None:
                break
            sec = sec.parent
            if "glove benefits" in sec.get_text(" ", strip=True).lower():
                break
        # take first 600 chars
        dtxt = sec.get_text(" ", strip=True)
        desc = dtxt[:600] if dtxt else None

    return {
        "title_detail": title,
        "price_detail": price,
        "model_code": model_code,
        "glove_profile": glove_profile,
        "description_snippet": desc,
        "images": images,
    }


def write_catalog_rows(ws, items: List[Dict]) -> Tuple[int, int]:
    hm = header_map(ws)
    required = ["product_id", "product_url"]
    for r in required:
        if r not in hm:
            raise RuntimeError(f"Sheet {ws.title} missing required header: {r}")

    idx = build_existing_index(ws, "product_id")
    appended = 0
    for it in items:
        pid = str(it["product_id"])
        if pid in idx:
            continue
        row = first_empty_row(ws, hm["product_id"])
        ws.cell(row=row, column=hm["product_id"]).value = pid
        ws.cell(row=row, column=hm["product_url"]).value = it["product_url"]
        if "source" in hm:
            ws.cell(row=row, column=hm["source"]).value = "JBG"
        if "catalog_title" in hm and it.get("title_catalog"):
            ws.cell(row=row, column=hm["catalog_title"]).value = it["title_catalog"]
        if "catalog_price" in hm and it.get("price_catalog") is not None:
            ws.cell(row=row, column=hm["catalog_price"]).value = float(it["price_catalog"])
        if "thumb_url" in hm and it.get("thumb_url"):
            ws.cell(row=row, column=hm["thumb_url"]).value = it["thumb_url"]
        if "catalog_scraped_at" in hm:
            ws.cell(row=row, column=hm["catalog_scraped_at"]).value = now_iso()
        appended += 1

    return len(items), appended


def collect_detail_targets(ws_catalog, ws_detail, resume: bool) -> List[Tuple[str, str]]:
    """
    Return [(product_id, product_url)] that need detail scrape.
    If resume=True, skip ones already present in detail sheet OR with detail_status=OK.
    """
    hmc = header_map(ws_catalog)
    hmd = header_map(ws_detail)

    for col in ("product_id", "product_url"):
        if col not in hmc:
            raise RuntimeError(f"Sheet {ws_catalog.title} missing required header: {col}")
    if "product_id" not in hmd:
        raise RuntimeError(f"Sheet {ws_detail.title} missing required header: product_id")

    existing_detail = build_existing_index(ws_detail, "product_id")

    targets = []
    for r in range(2, ws_catalog.max_row + 1):
        pid = ws_catalog.cell(row=r, column=hmc["product_id"]).value
        url = ws_catalog.cell(row=r, column=hmc["product_url"]).value
        if not pid or not url:
            continue
        pid = str(pid).strip()
        url = str(url).strip()

        if resume and pid in existing_detail:
            # if already scraped OK, skip
            drow = existing_detail[pid]
            status = ws_detail.cell(row=drow, column=hmd.get("detail_status", 0)).value if hmd.get("detail_status") else None
            if status and str(status).upper().startswith("OK"):
                continue
        targets.append((pid, url))
    return targets


def upsert_detail_row(ws_detail, pid: str, url: str, data: Dict, ok: bool, err: Optional[str] = None):
    hm = header_map(ws_detail)
    if "product_id" not in hm:
        raise RuntimeError(f"Sheet {ws_detail.title} missing required header: product_id")

    idx = build_existing_index(ws_detail, "product_id")
    if pid in idx:
        row = idx[pid]
    else:
        row = first_empty_row(ws_detail, hm["product_id"])
        ws_detail.cell(row=row, column=hm["product_id"]).value = pid

    # common fields
    if "product_url" in hm:
        ws_detail.cell(row=row, column=hm["product_url"]).value = url
    if "detail_scraped_at" in hm:
        ws_detail.cell(row=row, column=hm["detail_scraped_at"]).value = now_iso()
    if "detail_status" in hm:
        ws_detail.cell(row=row, column=hm["detail_status"]).value = "OK" if ok else "ERR"
    if "detail_error" in hm:
        ws_detail.cell(row=row, column=hm["detail_error"]).value = err

    # structured payloads
    if ok:
        if "title" in hm and data.get("title_detail"):
            ws_detail.cell(row=row, column=hm["title"]).value = data["title_detail"]
        if "price" in hm and data.get("price_detail") is not None:
            ws_detail.cell(row=row, column=hm["price"]).value = float(data["price_detail"])
        if "model_code" in hm and data.get("model_code"):
            ws_detail.cell(row=row, column=hm["model_code"]).value = data["model_code"]
        if "glove_profile_json" in hm:
            ws_detail.cell(row=row, column=hm["glove_profile_json"]).value = safe_json(data.get("glove_profile") or {})
        if "description_snippet" in hm and data.get("description_snippet"):
            ws_detail.cell(row=row, column=hm["description_snippet"]).value = data["description_snippet"]
        if "images_json" in hm:
            ws_detail.cell(row=row, column=hm["images_json"]).value = safe_json(data.get("images") or [])

        # Also store a single merged spec_json for Codex
        if "spec_json" in hm:
            merged = {
                "model_code": data.get("model_code"),
                "glove_profile": data.get("glove_profile") or {},
            }
            ws_detail.cell(row=row, column=hm["spec_json"]).value = safe_json(merged)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to GloveIQ master template XLSX")
    ap.add_argument("--start-url", default=DEFAULT_START_URL, help="Catalog start URL")
    ap.add_argument("--delay", type=float, default=1.25, help="Delay between requests (seconds)")
    ap.add_argument("--max-pages", type=int, default=25, help="Max catalog pages to crawl (0 to skip catalog phase)")
    ap.add_argument("--max-details", type=int, default=0, help="Max product detail pages to scrape (0 to skip details)")
    ap.add_argument("--resume", action="store_true", help="Skip already-scraped detail rows with detail_status=OK")
    args = ap.parse_args()

    sess = get_session()
    wb = load_wb(args.xlsx)

    if "JBG_Full_Catalog" not in wb.sheetnames or "JBG_Detail_Enrichment" not in wb.sheetnames:
        raise SystemExit("XLSX must contain sheets: JBG_Full_Catalog and JBG_Detail_Enrichment")

    ws_cat = wb["JBG_Full_Catalog"]
    ws_det = wb["JBG_Detail_Enrichment"]

    # -------------------
    # Catalog phase
    # -------------------
    if args.max_pages > 0:
        start = args.start_url
        print(f"[JBG CATALOG] Start: {start}")
        total_unique_before = len(build_existing_index(ws_cat, "product_id"))
        total_discovered = 0
        total_appended = 0

        page = 1
        url = start
        while page <= args.max_pages:
            if page == 1:
                url = start
            else:
                url = set_query_param(start, "page", str(page))
            print(f"[JBG CATALOG] Fetch page {page}: {url}")
            resp = sess.get(url, timeout=30)
            resp.raise_for_status()

            items = extract_catalog_products(resp.text, url)
            discovered, appended = write_catalog_rows(ws_cat, items)
            total_discovered += discovered
            total_appended += appended

            total_unique_now = len(build_existing_index(ws_cat, "product_id"))
            print(f"[JBG CATALOG] Page {page}: discovered={discovered} appended_new={appended} total_unique={total_unique_now}")

            # stop if no new appended (likely end of pagination / filtered list exhausted)
            if appended == 0 and page > 1:
                print("[JBG CATALOG] No new items appended; stop.")
                break

            time.sleep(args.delay)
            page += 1

        wb.save(args.xlsx)
        print(f"[JBG CATALOG] Saved workbook after catalog phase: {args.xlsx} (new={total_appended}, was={total_unique_before})")

    # -------------------
    # Detail phase
    # -------------------
    if args.max_details > 0:
        targets = collect_detail_targets(ws_cat, ws_det, resume=args.resume)
        if args.resume:
            # filter further: skip if detail already exists and is OK
            pass
        print(f"[JBG DETAIL] Targets: {len(targets)} (resume={args.resume})")
        count = 0
        for (pid, url) in targets:
            if count >= args.max_details:
                break
            count += 1
            print(f"[JBG DETAIL] ({count}/{min(len(targets), args.max_details)}) {pid}")
            try:
                resp = sess.get(url, timeout=30)
                resp.raise_for_status()
                data = parse_detail_page(resp.text, url)
                upsert_detail_row(ws_det, pid, url, data, ok=True, err=None)
            except Exception as e:
                upsert_detail_row(ws_det, pid, url, data={}, ok=False, err=str(e))
                print(f"[JBG DETAIL] Error on {pid}: {e}")

            # checkpoint every 50
            if count % 50 == 0:
                wb.save(args.xlsx)
                print(f"[JBG DETAIL] Checkpoint saved at {count} rows.")

            time.sleep(args.delay)

        wb.save(args.xlsx)
        print(f"[DONE] Final workbook saved: {args.xlsx}")


if __name__ == "__main__":
    main()
